
import io
import re
import zipfile
from pathlib import Path
from datetime import datetime
from copy import copy

import pdfplumber
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

STATE_CODES = {
    "01":"Jammu & Kashmir","02":"Himachal Pradesh","03":"Punjab","04":"Chandigarh",
    "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan","09":"Uttar Pradesh",
    "10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur",
    "15":"Mizoram","16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
    "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh",
    "24":"Gujarat","25":"Daman & Diu","26":"Dadra & Nagar Haveli and Daman & Diu",
    "27":"Maharashtra","28":"Andhra Pradesh (Old Code)","29":"Karnataka","30":"Goa",
    "31":"Lakshadweep","32":"Kerala","33":"Tamil Nadu","34":"Puducherry","35":"Andaman & Nicobar Islands",
    "36":"Telangana","37":"Andhra Pradesh","38":"Ladakh","97":"Other Territory","99":"Other",
}

def state_from_gstin(gstin):
    s = str(gstin or "").strip().upper()
    return STATE_CODES.get(s[:2], "Unknown State") if len(s) >= 2 else "Unknown State"

def is_nil_record(rec):
    """Return True only for a genuine NIL 3B, never merely because a PDF
    contains the word 'nil' (for example, 'nil-rated supplies' in Table 3.1).

    A NIL return is identified by an explicit NIL marker in the filename or
    by every extracted monetary field being zero. Generic PDF text is NOT
    inspected for the word NIL because normal GSTR-3B PDFs contain it often.
    """
    fn = Path(str(rec.get("file_name", ""))).name.lower()
    explicit_filename = bool(re.search(r"(?:^|[_\-\s])nil(?:[_\-\s.]|$)", fn))

    nums = []
    for x in rec.get("sections", []):
        nums += [x.get("taxable",0), x.get("igst",0), x.get("cgst",0), x.get("sgst",0), x.get("cess",0)]
    for x in rec.get("payment", []):
        nums += [x.get("tax_payable",0), x.get("negative_adjustment",0), x.get("net_tax_payable",0),
                 x.get("itc_igst",0), x.get("itc_cgst",0), x.get("itc_sgst",0), x.get("itc_cess",0),
                 x.get("cash_tax",0), x.get("interest_cash",0), x.get("late_fee_cash",0)]
    for x in rec.get("tax_breakup", []):
        nums += [x.get("igst",0), x.get("cgst",0), x.get("sgst",0), x.get("cess",0)]

    has_financial_fields = bool(nums)
    all_zero = has_financial_fields and all(abs(float(v or 0)) < 1e-9 for v in nums)
    return explicit_filename or all_zero


def clean_text(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def num(v):
    if v is None:
        return 0.0
    s = str(v).replace(",", "").replace("\n", "").replace(" ", "")
    s = s.replace("₹", "")
    # PDF extraction can produce values such as 7375884..00
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", ".", "-."):
        return 0.0
    # Keep only the last decimal point if extraction duplicated punctuation.
    if s.count(".") > 1:
        first = s.find(".")
        s = s[:first + 1] + s[first + 1:].replace(".", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def month_display(dt):
    return dt.strftime("%b-%y") if dt else ""


def extract_pdf(pdf_bytes):
    result = {
        "gstin": "",
        "legal_name": "",
        "trade_name": "",
        "arn": "",
        "arn_date": "",
        "fy": "",
        "tax_period": "",
        "month": None,
        "sections": [],
        "payment": [],
        "tax_breakup": [],
        "source_pages": {},
    }

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        # Metadata and tables
        for pno, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            result["source_pages"][pno] = []
            for ti, table in enumerate(tables):
                result["source_pages"][pno].append((ti, table))

        # Metadata: filed-copy PDFs can have different line wrapping depending
        # on how they were downloaded/printed. Search all pages and allow
        # flexible spaces, punctuation and line breaks.
        all_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        text = all_text
        result["_raw_text"] = text

        year_patterns = [
            r"\bYear\s*[:\-]?\s*(20\d{2}\s*[-/]\s*\d{2})\b",
            r"\bFY\s*[:\-]?\s*(20\d{2}\s*[-/]\s*\d{2})\b",
            r"\bFinancial\s+Year\s*[:\-]?\s*(20\d{2}\s*[-/]\s*\d{2})\b",
        ]
        for pat in year_patterns:
            m = re.search(pat, text, re.I)
            if m:
                result["fy"] = re.sub(r"\s+", "", m.group(1)).replace("/", "-")
                break

        period_patterns = [
            r"\bPeriod\s*[:\-]?\s*([A-Za-z]{3,12})\b",
            r"\bTax\s+Period\s*[:\-]?\s*([A-Za-z]{3,12})\b",
        ]
        for pat in period_patterns:
            m = re.search(pat, text, re.I)
            if m:
                candidate = m.group(1).strip().title()
                if candidate.lower() in MONTHS:
                    result["tax_period"] = candidate
                    break

        # Fallback: infer FY/month from common filed-copy filenames such as
        # GSTR3B_29AAACI5754J1Z8_042025.pdf or ..._April_2025.pdf.
        if not result["tax_period"] or not result["fy"]:
            fname = getattr(pdf.stream, "name", "") or ""
            name = Path(str(fname)).name
            m = re.search(r"(0[1-9]|1[0-2])\s*(20\d{2})", name)
            if m:
                mm, yy = int(m.group(1)), int(m.group(2))
                result["tax_period"] = list(MONTHS.keys())[mm - 1].title()
                start = yy if mm >= 4 else yy - 1
                result["fy"] = f"{start}-{str(start + 1)[-2:]}"
            else:
                m = re.search(
                    r"(January|February|March|April|May|June|July|August|September|October|November|December)[_\- ]*(20\d{2})",
                    name, re.I
                )
                if m:
                    month_name = m.group(1).title()
                    yy = int(m.group(2))
                    mm = MONTHS[month_name.lower()]
                    result["tax_period"] = month_name
                    start = yy if mm >= 4 else yy - 1
                    result["fy"] = f"{start}-{str(start + 1)[-2:]}"
        m = re.search(r"GSTIN of the supplier\s+([0-9A-Z]+)", text)
        if m:
            result["gstin"] = m.group(1)
        m = re.search(r"2\(a\)\.\s*Legal name of the registered person\s+(.+)", text)
        if m:
            result["legal_name"] = clean_text(m.group(1))
        m = re.search(r"2\(b\)\.\s*Trade name, if any\s+(.+)", text)
        if m:
            result["trade_name"] = clean_text(m.group(1))
        m = re.search(r"2\(c\)\.\s*ARN\s+(.+)", text)
        if m:
            result["arn"] = clean_text(m.group(1))
        m = re.search(r"2\(d\)\.\s*Date of ARN\s+(.+)", text)
        if m:
            result["arn_date"] = clean_text(m.group(1))

        if result["fy"] and result["tax_period"].lower() in MONTHS:
            fy_start = int(result["fy"][:4])
            fy_end = int("20" + result["fy"][-2:])
            y = fy_start if MONTHS[result["tax_period"].lower()] >= 4 else fy_end
            result["month"] = datetime(y, MONTHS[result["tax_period"].lower()], 1)

        def add(section, head, taxable=0, igst=0, cgst=0, sgst=0, cess=0, page=None):
            result["sections"].append({
                "section": section,
                "head": head,
                "taxable": taxable,
                "igst": igst,
                "cgst": cgst,
                "sgst": sgst,
                "cess": cess,
                "page": page,
            })

        # 3.1 and 3.1.1
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for pno, page in enumerate(pdf.pages, 1):
                for table in page.extract_tables():
                    if not table:
                        continue
                    headers = " ".join(clean_text(x) for x in (table[0] or []))
                    if "Total taxable" in headers and "Integrated" in headers and "Cess" in headers:
                        for row in table[1:]:
                            if not row:
                                continue
                            label = clean_text(row[0])
                            if not label or label in ("A. ITC Available (whether in full or part)", "B. ITC Reversed"):
                                continue
                            vals = [num(x) for x in row[1:]]
                            if len(vals) >= 5:
                                if label.startswith("(a)") or label.startswith("(b)") or label.startswith("(c)") or label.startswith("(d)") or label.startswith("(e)"):
                                    add("3.1", label, *vals[:5], page=pno)
                                elif "(i)" in label or "(ii)" in label:
                                    add("3.1.1", label, *vals[:5], page=pno)
                    # 4 ITC (page 2 may continue without repeating the header)
                    itc_table = (
                        headers.startswith("Details") and "Integrated tax" in headers and "Cess" in headers
                    ) or any(
                        "ITC" in clean_text(row[0]) or "rules 38,42" in clean_text(row[0]).lower()
                        for row in table if row and row[0]
                    )
                    if itc_table:
                        start_rows = table[1:] if headers.startswith("Details") else table
                        for row in start_rows:
                            if not row:
                                continue
                            label = clean_text(row[0])
                            vals = [num(x) for x in row[1:]]
                            if len(vals) >= 4 and (
                                re.match(r"^\([1-5]\)", label)
                                or label.startswith("C. Net ITC")
                                or label.startswith("(D)")
                            ):
                                # Determine page/section based on labels.
                                if any(k in label.lower() for k in ["import of goods", "import of services", "inward supplies", "inward supplies from isd", "all other itc"]):
                                    sec = "4A - ITC Available"
                                elif "rules 38,42" in label.lower() or label.lower() == "(2) others":
                                    sec = "4B - ITC Reversed"
                                elif "net itc available" in label.lower():
                                    sec = "4C - Net ITC Available"
                                elif label.startswith("(D)") or "itc reclaimed" in label.lower() or "ineligible itc" in label.lower():
                                    sec = "4D - Other Details"
                                else:
                                    sec = "4 - Other"
                                add(sec, label, 0, *vals[:4], page=pno)

                    # 3.2
                    if headers == "Nature of Supplies Total taxable value Integrated tax":
                        for row in table[1:]:
                            if len(row) >= 3:
                                add("3.2", clean_text(row[0]), num(row[1]), num(row[2]), 0, 0, 0, page=pno)

                    # 5
                    if "Inter- State supplies" in headers and "Intra- State supplies" in headers:
                        for row in table[1:]:
                            if len(row) >= 3:
                                add("5", clean_text(row[0]), num(row[1]), num(row[2]), 0, 0, 0, page=pno)

                    # 5.1
                    if headers.startswith("Details Integrated tax") and "Late fee" in headers:
                        for row in table[1:]:
                            if len(row) >= 5:
                                label = clean_text(row[0])
                                if label:
                                    add("5.1", label, 0, num(row[1]), num(row[2]), num(row[3]), num(row[4]), page=pno)

                    # Payment 6.1
                    if "Net Tax" in headers and "Tax paid" in headers and "Interest" in headers:
                        mode = ""
                        for row in table[2:]:
                            if not row:
                                continue
                            label = clean_text(row[0])
                            if "Other than reverse charge" in label:
                                mode = "6.1(A) Other than reverse charge"
                                continue
                            if "Reverse charge and supplies made" in label:
                                mode = "6.1(B) Reverse charge and supplies made u/s 9(5)"
                                continue
                            if label in ("Integrated tax", "Central tax", "State/UT tax", "Cess"):
                                vals = [num(x) for x in row[1:]]
                                while len(vals) < 10:
                                    vals.append(0.0)
                                result["payment"].append({
                                    "section": mode,
                                    "head": label,
                                    "tax_payable": vals[0],
                                    "negative_adjustment": vals[1],
                                    "net_tax_payable": vals[2],
                                    "itc_igst": vals[3],
                                    "itc_cgst": vals[4],
                                    "itc_sgst": vals[5],
                                    "itc_cess": vals[6],
                                    "cash_tax": vals[7],
                                    "interest_cash": vals[8],
                                    "late_fee_cash": vals[9],
                                    "page": pno,
                                })

                    # Tax liability breakup: header is on page 2 and data can continue on page 3.
                    if "Period" in headers and "Integrated tax" in headers and "Central tax" in headers and "State/UT tax" in headers and "Cess" in headers:
                        for row in table[1:]:
                            if len(row) >= 5:
                                result["tax_breakup"].append({
                                    "period": clean_text(row[0]),
                                    "igst": num(row[1]),
                                    "cgst": num(row[2]),
                                    "sgst": num(row[3]),
                                    "cess": num(row[4]),
                                    "page": pno,
                                })
                    elif len(table) == 1 and len(table[0]) >= 5:
                        first = clean_text(table[0][0])
                        if re.search(r"\b(20\d{2})\b", first) and re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)", first, re.I):
                            result["tax_breakup"].append({
                                "period": first,
                                "igst": num(table[0][1]),
                                "cgst": num(table[0][2]),
                                "sgst": num(table[0][3]),
                                "cess": num(table[0][4]),
                                "page": pno,
                            })

        return result



def _copy_row_style(ws, source_row, target_row):
    for c in range(1, ws.max_column + 1):
        src = ws.cell(source_row, c)
        dst = ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _prepare_data_rows(ws, n_records, template_data_row=4, template_total_row=5):
    # The supplied template has one sample data row followed by Grand Total.
    # Expand it while preserving all formatting.
    if n_records <= 0:
        n_records = 1
    if n_records > 1:
        ws.insert_rows(template_total_row, amount=n_records - 1)
        for r in range(template_data_row + 1, template_data_row + n_records):
            _copy_row_style(ws, template_data_row, r)
    total_row = template_data_row + n_records
    # Clear all data cells; labels/merged headers remain untouched.
    for r in range(template_data_row, total_row):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
    ws.cell(total_row, 1, "Grand Total")
    return total_row


def _write_month_rows(ws, records, row_writer, total_row, numeric_cols):
    for idx, rec in enumerate(records):
        row = 4 + idx
        row_writer(row, rec)
        for c in numeric_cols:
            ws.cell(row, c).number_format = '#,##0.00'
    for c in numeric_cols:
        letter = get_column_letter(c)
        ws.cell(total_row, c, f"=SUM({letter}4:{letter}{total_row-1})")
        ws.cell(total_row, c).number_format = '#,##0.00'
    for c in range(1, ws.max_column + 1):
        ws.cell(total_row, c).fill = PatternFill("solid", fgColor="F3E6EB")
        ws.cell(total_row, c).font = Font(bold=True, color="98002E")
        ws.cell(total_row, c).border = Border(bottom=Side(style="thin", color="D9DCE3"))




def consolidate_state_month_records(records):
    """Consolidate multiple GSTIN records into one record per State + Month."""
    grouped = {}
    for rec in records:
        state = rec.get("state") or state_from_gstin(rec.get("gstin")) or "Unknown State"
        key = (state, rec.get("month"))
        if key not in grouped:
            base = dict(rec)
            base["state"] = state
            base["gstin"] = "Multiple GSTINs"
            base["sections"] = []
            base["payment"] = []
            grouped[key] = base
        out = grouped[key]
        # Merge section rows by section + particular/head.
        secmap = {(x.get("section"), clean_text(x.get("head"))): x for x in out["sections"]}
        for x in rec.get("sections", []):
            k = (x.get("section"), clean_text(x.get("head")))
            if k not in secmap:
                nx = dict(x)
                out["sections"].append(nx)
                secmap[k] = nx
            else:
                nx = secmap[k]
                for fld in ("taxable", "igst", "cgst", "sgst", "cess"):
                    nx[fld] = (nx.get(fld) or 0) + (x.get(fld) or 0)
        # Merge payment rows by section + tax head.
        paymap = {(x.get("section"), clean_text(x.get("head"))): x for x in out["payment"]}
        for x in rec.get("payment", []):
            k = (x.get("section"), clean_text(x.get("head")))
            if k not in paymap:
                nx = dict(x)
                out["payment"].append(nx)
                paymap[k] = nx
            else:
                nx = paymap[k]
                for fld in ("tax_payable", "negative_adjustment", "net_tax_payable", "itc_igst", "itc_cgst", "itc_sgst", "itc_cess", "cash_tax", "interest_cash", "late_fee_cash"):
                    nx[fld] = (nx.get(fld) or 0) + (x.get(fld) or 0)
        statuses = {str(out.get("filing_status") or "FILED").upper(), str(rec.get("filing_status") or "FILED").upper()}
        out["filing_status"] = "FILED" if "FILED" in statuses else ("NIL FILED" if "NIL FILED" in statuses or "NIL" in statuses else next(iter(statuses)))
    # Preserve the number of GST registrations represented by each state.
    state_gstins = {}
    for rec in records:
        state = rec.get("state") or state_from_gstin(rec.get("gstin")) or "Unknown State"
        state_gstins.setdefault(state, set()).add(str(rec.get("gstin") or "").strip())
    for out in grouped.values():
        state = out.get("state") or "Unknown State"
        valid = {g for g in state_gstins.get(state, set()) if g}
        out["registration_count"] = len(valid)
        out["registration_gstins"] = sorted(valid)
    return [grouped[k] for k in sorted(grouped, key=lambda k: (k[0], k[1] or datetime.max))]


def _write_statewise_month_rows(ws, records, row_writer, numeric_cols, label_cols=None):
    """Write every analytical sheet in the same visual structure:
    State header (with registration count) -> one row per month -> State Total
    -> blank row -> next state -> Grand Total.
    """
    if not records:
        return 4
    states = {}
    for rec in records:
        state = rec.get("state") or state_from_gstin(rec.get("gstin")) or "Unknown State"
        states.setdefault(state, []).append(rec)

    # Rebuild the data area from row 4 onward so the template header remains intact.
    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row - 3)

    row = 4
    grand = {c: 0.0 for c in numeric_cols}
    for state in sorted(states):
        recs = sorted(states[state], key=lambda r: r.get("month") or datetime.max)
        reg_count = len({str(r.get("gstin") or "").strip() for r in recs if r.get("gstin") and str(r.get("gstin")).strip()})
        # State header exactly in the requested style.
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ws.max_column)
        cell = ws.cell(row, 1, f"{state}  •  {reg_count} registration(s)")
        cell.fill = PatternFill("solid", fgColor="F3E6EB")
        cell.font = Font(name="Trebuchet MS", size=10, bold=True, color="98002E")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        state_numeric = {c: 0.0 for c in numeric_cols}
        for rec in recs:
            row_writer(row, rec)
            for c in numeric_cols:
                v = ws.cell(row, c).value
                if isinstance(v, (int, float)):
                    state_numeric[c] += float(v)
                    grand[c] += float(v)
                ws.cell(row, c).number_format = "#,##0.00"
            row += 1

        ws.cell(row, 1, f"{state} Total")
        for c in numeric_cols:
            ws.cell(row, c, state_numeric[c])
            ws.cell(row, c).number_format = "#,##0.00"
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row, c)
            cell.fill = PatternFill("solid", fgColor="F3E6EB")
            cell.font = Font(name="Trebuchet MS", size=10, bold=True, color="98002E")
            cell.border = Border(bottom=Side(style="thin", color="D9DCE3"))
        row += 2  # blank separator between states

    total_row = row
    ws.cell(total_row, 1, "Grand Total")
    for c in numeric_cols:
        ws.cell(total_row, c, grand[c])
        ws.cell(total_row, c).number_format = "#,##0.00"
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(total_row, c)
        cell.fill = PatternFill("solid", fgColor="98002E")
        cell.font = Font(name="Trebuchet MS", size=10, bold=True, color="FFFFFF")
    return total_row


def build_workbook(records):
    """
    Build the workbook from the user's supplied Excel template.

    IMPORTANT: The template controls the presentation. Values are populated
    from the filed GSTR-3B PDF parser field-by-field; no cross-section
    substitution is performed.
    """
    records = sorted(records, key=lambda r: r["month"] or datetime.max)
    state_month_records = consolidate_state_month_records(records)

    template_path = Path(__file__).with_name("GSTR3B_Output_Template.xlsx")
    if template_path.exists():
        wb = load_workbook(template_path)
    else:
        # Fallback for environments where the template is unavailable.
        wb = Workbook()

    maroon = "98002E"
    light = "F3E6EB"
    thin = Side(style="thin", color="D9DCE3")

    # Remove generated/obsolete sheets if present; then normalize the
    # workbook around the user's supplied layout.
    for sn in ["GSTR-3B Summary", "4. Eligible ITC", "Reconciliation"]:
        if sn in wb.sheetnames:
            del wb[sn]

    # ------------------------------------------------------------------
    # Read Me
    # ------------------------------------------------------------------
    if "Read Me" in wb.sheetnames:
        wm = wb["Read Me"]
    else:
        wm = wb.create_sheet("Read Me", 0)
    # --------------------------------------------------------------
    # Professional Read Me / branding page
    # --------------------------------------------------------------
    gstins = sorted({str(r.get("gstin", "")).strip() for r in records if r.get("gstin")})
    gstin_text = ", ".join(gstins) if gstins else "Not detected"

    wm["B1"] = "GSTR-3B PDF → Excel Analyzer"
    wm["B2"] = "Designed & Developed by Pushpak Kumar"
    wm["B3"] = "Website: pushpakkumar.com"
    wm["B4"] = "GSTIN"
    wm["C4"] = gstin_text
    wm["B5"] = "Financial Year"
    wm["C5"] = ", ".join(sorted({str(r.get("fy", "")).strip() for r in records if r.get("fy")})) or "Not detected"
    wm["B6"] = "Tax Periods"
    wm["C6"] = ", ".join(month_display(r.get("month")) for r in records if r.get("month")) or "Not detected"
    wm["B7"] = "Purpose"
    wm["C7"] = "Month-wise extraction and consolidation of filed GSTR-3B PDF copies."
    wm["B8"] = "Coverage"
    wm["C8"] = "Tables 3.1, 3.1.1, 3.2, 4, 5, 5.1, 6.1 and tax-liability breakup."
    wm["B9"] = "Mapping"
    wm["C9"] = "Values are mapped field-by-field from the filed PDF. Source pages are retained in Extraction Audit."
    wm["B11"] = "BRANDING"
    wm["B12"] = "Pushpak Kumar"
    wm["B13"] = "pushpakkumar.com"
    wm["B14"] = "GSTR-3B Analysis & Consolidation Tool"

    # Professional branding styles.
    maroon_fill = PatternFill("solid", fgColor="98002E")
    light_fill = PatternFill("solid", fgColor="F3E6EB")
    dark_font = Font(name="Calibri", size=12, bold=True, color="98002E")
    white_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=22, bold=True, color="FFFFFF")
    normal_font = Font(name="Calibri", size=11, color="333333")

    for cell in ["B1"]:
        wm[cell].fill = maroon_fill
        wm[cell].font = title_font
        wm[cell].alignment = Alignment(vertical="center")
    wm.merge_cells("B1:H1")
    wm.row_dimensions[1].height = 38

    for cell in ["B2", "B3", "B4", "B5", "B6", "B7", "B8", "B9"]:
        wm[cell].font = dark_font
        wm[cell].fill = light_fill
        wm[cell].alignment = Alignment(vertical="top", wrap_text=True)

    for cell in ["C4", "C5", "C6", "C7", "C8", "C9"]:
        wm[cell].font = normal_font
        wm[cell].alignment = Alignment(vertical="top", wrap_text=True)

    wm["B11"].fill = maroon_fill
    wm["B11"].font = white_font
    wm.merge_cells("B11:H11")
    wm["B12"].font = Font(name="Calibri", size=16, bold=True, color="98002E")
    wm["B13"].font = Font(name="Calibri", size=12, bold=True, color="98002E")
    wm["B14"].font = Font(name="Calibri", size=11, italic=True, color="666666")

    wm.column_dimensions["A"].width = 4
    wm.column_dimensions["B"].width = 24
    wm.column_dimensions["C"].width = 42
    for col in ["D", "E", "F", "G", "H"]:
        wm.column_dimensions[col].width = 18

    wm.freeze_panes = "B4"

    # ------------------------------------------------------------------
    # 3.1 Outward Details — exact user template format
    # ------------------------------------------------------------------
    ws = wb["3.1 Outward Details"]
    section_heads = [
        ("(a)", 2), ("(b)", 7), ("(c)", 12), ("(d)", 17), ("(e)", 22)
    ]
    def write31(row, rec):
        ws.cell(row, 1, month_display(rec["month"]))
        for label, col in section_heads:
            matches = [x for x in rec["sections"] if x["section"] == "3.1" and x["head"].lower().startswith(label.lower())]
            x = matches[0] if matches else {"taxable":0,"igst":0,"cgst":0,"sgst":0,"cess":0}
            vals = [x["taxable"], x["igst"], x["cgst"], x["sgst"], x["cess"]]
            for off, v in enumerate(vals):
                ws.cell(row, col + off, v)
    total_row = _write_statewise_month_rows(ws, state_month_records, write31, list(range(2,27)))
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:Z{total_row}"

    # ------------------------------------------------------------------
    # 3.1.1 Supplies under 9(5) — use the user's 11-column format
    # ------------------------------------------------------------------
    ws = wb["3.1.1   Supplies under 9(5)"]
    # The supplied workbook's 7-column sheet is actually 3.2 format.
    # Replace its contents with the 11-column 3.1.1 layout by copying the
    # user's existing 3.1.2 sheet, preserving that exact visual style.
    old = ws
    src = wb["3.1.2  Out of supplies made 3.1"]
    # Delete old sheet and clone source; openpyxl cannot reliably clone across
    # names with all objects, so use copy_worksheet.
    del wb["3.1.1   Supplies under 9(5)"]
    ws = wb.copy_worksheet(src)
    ws.title = "3.1.1 Supplies u-s 9(5)"
    # Keep source sheet as the 3.2 layout for now; it will be renamed below.
    ws["A1"] = "Month"
    ws["B1"] = "3.1.1 Details of Supplies notified under section 9(5) of the CGST Act, 2017 and corresponding provisions in IGST/UTGST/SGST Acts"
    ws["B2"] = "(i) Taxable supplies on which electronic commerce operator pays tax u/s 9(5) [to be furnished by electronic commerce operator]"
    ws["G2"] = "(ii) Taxable supplies made by registered person through electronic commerce operator, on which electronic commerce operator is required to pay tax u/s 9(5) [to be furnished by registered person making supplies through electronic commerce operator]"
    def write311(row, rec):
        ws.cell(row,1,month_display(rec["month"]))
        for label, col in [("(i)",2),("(ii)",7)]:
            matches=[x for x in rec["sections"] if x["section"]=="3.1.1" and re.search(rf"\({re.escape(label.strip('()'))}\)", x["head"], re.I)]
            x=matches[0] if matches else {"taxable":0,"igst":0,"cgst":0,"sgst":0,"cess":0}
            for off,v in enumerate([x["taxable"],x["igst"],x["cgst"],x["sgst"],x["cess"]]):
                ws.cell(row,col+off,v)
    total_row_311 = _write_statewise_month_rows(ws, state_month_records, write311, list(range(2,12)))
    ws.freeze_panes="B4"
    ws.auto_filter.ref=f"A3:K{total_row_311}"

    # ------------------------------------------------------------------
    # 3.2 — use the user's 7-column format
    # ------------------------------------------------------------------
    ws32 = wb["3.1.2  Out of supplies made 3.1"]
    # Remove the original 11-column merges before rebuilding the 7-column
    # 3.2 presentation.
    for mr in list(ws32.merged_cells.ranges):
        ws32.unmerge_cells(str(mr))
    ws32.title = "3.2 Inter-State Supplies"
    ws32["B1"] = "3.2 Out of supplies made in 3.1(a) and 3.1.1(i), details of inter-state supplies made"
    ws32["B2"] = "Supplies made to Unregistered Persons"
    ws32["D2"] = "Supplies made to Composition Taxable Persons"
    ws32["F2"] = "Supplies made to UIN holders"
    ws32["C3"] = "Integrated tax"
    ws32["E3"] = "Integrated tax"
    ws32["G3"] = "Integrated tax"
    # We need the 7-column layout. Rebuild from the original 7-column sheet
    # style stored in the user's workbook via the current sheet's columns.
    # The existing workbook was originally 11 columns; reconstruct the 7
    # column presentation using the same maroon/light style.
    for row in range(1, 6):
        for col in range(1, 12):
            if col > 7:
                ws32.cell(row,col).value=None
    ws32.merge_cells("A1:A3")
    ws32.merge_cells("B1:G1")
    ws32.merge_cells("B2:C2")
    ws32.merge_cells("D2:E2")
    ws32.merge_cells("F2:G2")
    for cell in ws32[1]:
        if cell.column <= 7:
            cell.fill=PatternFill("solid", fgColor=maroon)
            cell.font=Font(color="FFFFFF",bold=True)
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for c in [2,4,6]:
        ws32.cell(2,c).fill=PatternFill("solid",fgColor=light)
        ws32.cell(2,c).font=Font(color=maroon,bold=True)
        ws32.cell(2,c).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for c in [3,5,7]:
        ws32.cell(3,c).value="Integrated tax"
        ws32.cell(3,c).fill=PatternFill("solid",fgColor=maroon)
        ws32.cell(3,c).font=Font(color="FFFFFF",bold=True)
        ws32.cell(3,c).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws32.cell(3,2,"Total Taxable Value")
    ws32.cell(3,4,"Total Taxable Value")
    ws32.cell(3,6,"Total Taxable Value")
    def write32(row, rec):
        ws32.cell(row,1,month_display(rec["month"]))
        mappings=[
            ("Supplies made to Unregistered Persons",2),
            ("Supplies made to Composition Taxable Persons",4),
            ("Supplies made to UIN holders",6),
        ]
        for label,col in mappings:
            matches=[x for x in rec["sections"] if x["section"]=="3.2" and clean_text(x["head"]).lower()==label.lower()]
            x=matches[0] if matches else {"taxable":0,"igst":0}
            ws32.cell(row,col,x["taxable"])
            ws32.cell(row,col+1,x["igst"])
    total_row_32 = _write_statewise_month_rows(ws32, state_month_records, write32, list(range(2,8)))
    ws32.freeze_panes="B4"
    ws32.auto_filter.ref=f"A3:G{total_row_32}"
    for c,w in enumerate([13,22,16,22,16,18,16],1):
        ws32.column_dimensions[get_column_letter(c)].width=w

    # ------------------------------------------------------------------
    # 4. ITC Details — EXACT supplied format, no extra ITC sheet.
    # ------------------------------------------------------------------
    wi = wb["4. ITC Details"]
    itc_map = [
        ("4A - ITC Available","(1) Import of goods",2),
        ("4A - ITC Available","(2) Import of services",6),
        ("4A - ITC Available","(3) Inward supplies liable to reverse charge (other than 1 & 2 above)",10),
        ("4A - ITC Available","(4) Inward supplies from ISD",14),
        ("4A - ITC Available","(5) All other ITC",18),
        ("4B - ITC Reversed","(1) As per rules 38,42 & 43 of CGST Rules and section 17(5)",22),
        ("4B - ITC Reversed","(2) Others",26),
        ("4C - Net ITC Available","C. Net ITC available (A-B)",30),
        ("4D - Other Details","(1) ITC reclaimed which was reversed under Table 4(B)(2) in earlier tax period",34),
        ("4D - Other Details","(2) Ineligible ITC under section 16(4) & ITC restricted due to PoS rules",38),
    ]
    def find_itc(rec, sec, head):
        rows=[x for x in rec["sections"] if x["section"]==sec and clean_text(x["head"]).lower()==clean_text(head).lower()]
        if not rows and "inward supplies" in head.lower():
            rows=[x for x in rec["sections"] if x["section"]==sec and "inward supplies" in clean_text(x["head"]).lower() and (
                ("reverse charge" in head.lower()) == ("reverse charge" in x["head"].lower())
            )]
        if not rows:
            rows=[x for x in rec["sections"] if x["section"]==sec and clean_text(head).lower() in clean_text(x["head"]).lower()]
        return rows[0] if rows else {"igst":0,"cgst":0,"sgst":0,"cess":0}
    def write4(row, rec):
        wi.cell(row,1,month_display(rec["month"]))
        for sec,head,col in itc_map:
            x=find_itc(rec,sec,head)
            for off,v in enumerate([x["igst"],x["cgst"],x["sgst"],x["cess"]]):
                wi.cell(row,col+off,v)
    total_row_4 = _write_statewise_month_rows(wi, state_month_records, write4, list(range(2,42)))
    wi.freeze_panes="B4"
    wi.auto_filter.ref=f"A3:AO{total_row_4}"

    # ------------------------------------------------------------------
    # 5. Values of exempt, nil-rated and non-GST inward supplies
    # ------------------------------------------------------------------
    if "5. Inward Supplies" in wb.sheetnames:
        ws5=wb["5. Inward Supplies"]
    else:
        ws5=wb.create_sheet("5. Inward Supplies")
    ws5.delete_rows(1, ws5.max_row)
    headers=["Month","Particular","Inter-State supplies","Intra-State supplies","Page"]
    for c,h in enumerate(headers,1):
        cell=ws5.cell(1,c,h); cell.fill=PatternFill("solid",fgColor=maroon); cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",wrap_text=True)
    rr=2
    for state in sorted({r.get("state") for r in state_month_records}):
        state_recs=[r for r in state_month_records if r.get("state")==state]
        reg_count=sum(1 for r in records if r.get("state")==state and r.get("gstin"))
        reg_count=len({str(r.get("gstin")).strip() for r in records if r.get("state")==state and r.get("gstin")})
        ws5.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=5); ws5.cell(rr,1,f"{state}  •  {reg_count} registration(s)"); ws5.cell(rr,1).fill=PatternFill("solid",fgColor=light); ws5.cell(rr,1).font=Font(name="Trebuchet MS",size=10,bold=True,color=maroon); rr+=1
        for rec in state_recs:
            for x in [z for z in rec["sections"] if z["section"]=="5"]:
                ws5.cell(rr,1,month_display(rec["month"])); ws5.cell(rr,2,x["head"]); ws5.cell(rr,3,x["taxable"]); ws5.cell(rr,4,x["igst"]); ws5.cell(rr,5,x["page"]); rr+=1
        rr += 1
    ws5.freeze_panes="A2"; ws5.auto_filter.ref=ws5.dimensions
    for c,w in enumerate([13,70,20,20,8],1): ws5.column_dimensions[get_column_letter(c)].width=w
    for r in range(2,rr):
        for c in [3,4]: ws5.cell(r,c).number_format="#,##0.00"

    # ------------------------------------------------------------------
    # 5.1 Interest and Late Fee
    # ------------------------------------------------------------------
    if "5.1 Interest & Late Fee" in wb.sheetnames:
        ws51=wb["5.1 Interest & Late Fee"]
    else:
        ws51=wb.create_sheet("5.1 Interest & Late Fee")
    ws51.delete_rows(1, ws51.max_row)
    headers=["Month","Details","Integrated tax","Central tax","State/UT tax","Cess","Page"]
    for c,h in enumerate(headers,1):
        cell=ws51.cell(1,c,h); cell.fill=PatternFill("solid",fgColor=maroon); cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",wrap_text=True)
    rr=2
    for state in sorted({r.get("state") for r in state_month_records}):
        state_recs=[r for r in state_month_records if r.get("state")==state]
        reg_count=len({str(r.get("gstin")).strip() for r in records if r.get("state")==state and r.get("gstin")})
        ws51.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=7); ws51.cell(rr,1,f"{state}  •  {reg_count} registration(s)"); ws51.cell(rr,1).fill=PatternFill("solid",fgColor=light); ws51.cell(rr,1).font=Font(name="Trebuchet MS",size=10,bold=True,color=maroon); rr+=1
        for rec in state_recs:
            for x in [z for z in rec["sections"] if z["section"]=="5.1" and clean_text(z["head"]).lower() not in {"system computed -"}]:
                ws51.cell(rr,1,month_display(rec["month"])); ws51.cell(rr,2,x["head"]); ws51.cell(rr,3,x["igst"]); ws51.cell(rr,4,x["cgst"]); ws51.cell(rr,5,x["sgst"]); ws51.cell(rr,6,x["cess"]); ws51.cell(rr,7,x["page"]); rr+=1
        rr += 1
    ws51.freeze_panes="A2"; ws51.auto_filter.ref=ws51.dimensions
    for c,w in enumerate([13,25,18,18,18,15,8],1): ws51.column_dimensions[get_column_letter(c)].width=w
    for r in range(2,rr):
        for c in range(3,7): ws51.cell(r,c).number_format="#,##0.00"

    # ------------------------------------------------------------------
    # 6.1 Payment of Tax — supplied detailed format
    # ------------------------------------------------------------------
    wp = wb["6.1 Payment of Tax"]
    # Recreate this sheet cleanly while keeping the template's header style.
    wp.delete_rows(2, wp.max_row)
    pheaders=["Month","GSTIN","FY","Section","Tax Head","Tax Payable","Negative Liability Adjustment","Net Tax Payable","ITC IGST","ITC CGST","ITC SGST","ITC Cess","Tax Paid in Cash","Interest Paid in Cash","Late Fee Paid in Cash","Page"]
    for c,h in enumerate(pheaders,1):
        cell=wp.cell(1,c,h); cell.fill=PatternFill("solid",fgColor=maroon); cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",wrap_text=True)
    rr=2
    for state in sorted({r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State" for r in records}):
        reg_count=len({str(r.get("gstin")).strip() for r in records if (r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State")==state and r.get("gstin")})
        wp.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=16); wp.cell(rr,1,f"{state}  •  {reg_count} registration(s)"); wp.cell(rr,1).fill=PatternFill("solid",fgColor=light); wp.cell(rr,1).font=Font(name="Trebuchet MS",size=10,bold=True,color=maroon); rr+=1
        state_recs=sorted([r for r in records if (r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State")==state], key=lambda r:r.get("month") or datetime.max)
        for rec in state_recs:
            for x in rec["payment"]:
                vals=[month_display(rec["month"]),rec["gstin"],rec["fy"],x["section"],x["head"],x["tax_payable"],x["negative_adjustment"],x["net_tax_payable"],x["itc_igst"],x["itc_cgst"],x["itc_sgst"],x["itc_cess"],x["cash_tax"],x["interest_cash"],x["late_fee_cash"],x["page"]]
                for c,v in enumerate(vals,1): wp.cell(rr,c,v)
                rr+=1
        rr += 1
    for r in range(2,rr):
        for c in range(6,16): wp.cell(r,c).number_format="#,##0.00"
    wp.freeze_panes="A2"; wp.auto_filter.ref=wp.dimensions
    for c,w in enumerate([13,20,12,38,18,17,24,17,15,15,15,15,20,20,20,8],1): wp.column_dimensions[get_column_letter(c)].width=w

    # ------------------------------------------------------------------
    # Detailed — retain all sections, including ITC, so audit is complete.
    # ------------------------------------------------------------------
    if "GSTR-3B Detailed" in wb.sheetnames:
        wd=wb["GSTR-3B Detailed"]
        wd.delete_rows(2, wd.max_row)
    else:
        wd=wb.create_sheet("GSTR-3B Detailed")
    detail_headers=["State","Month","GSTIN","FY","Section","Particular","Taxable Value","IGST","CGST","SGST/UTGST","Cess","Page"]
    for c,h in enumerate(detail_headers,1):
        cell=wd.cell(1,c,h); cell.fill=PatternFill("solid",fgColor=maroon); cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",wrap_text=True)
    rr=2
    for state in sorted({r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State" for r in records}):
        reg_count=len({str(r.get("gstin")).strip() for r in records if (r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State")==state and r.get("gstin")})
        wd.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=12); wd.cell(rr,1,f"{state}  •  {reg_count} registration(s)"); wd.cell(rr,1).fill=PatternFill("solid",fgColor=light); wd.cell(rr,1).font=Font(name="Trebuchet MS",size=10,bold=True,color=maroon); rr+=1
        state_recs=sorted([r for r in records if (r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State")==state], key=lambda r:r.get("month") or datetime.max)
        for rec in state_recs:
            for x in rec["sections"]:
                vals=[state,month_display(rec["month"]),rec["gstin"],rec["fy"],x["section"],x["head"],x["taxable"],x["igst"],x["cgst"],x["sgst"],x["cess"],x["page"]]
                for c,v in enumerate(vals,1): wd.cell(rr,c,v)
                rr+=1
        rr += 1
    for r in range(2,rr):
        for c in range(7,12): wd.cell(r,c).number_format="#,##0.00"
    wd.freeze_panes="A2"; wd.auto_filter.ref=wd.dimensions
    for c,w in enumerate([22,13,20,12,28,72,18,16,16,16,16,8],1): wd.column_dimensions[get_column_letter(c)].width=w

    # ------------------------------------------------------------------
    # Extraction Audit
    # ------------------------------------------------------------------
    wa=wb["Extraction Audit"]
    wa.delete_rows(2,wa.max_row)
    audit_headers=["State","Month","GSTIN","FY","Source PDF","Page","Section","Particular","Value / Details"]
    for c,h in enumerate(audit_headers,1):
        cell=wa.cell(1,c,h); cell.fill=PatternFill("solid",fgColor=maroon); cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",wrap_text=True)
    rr=2
    for state in sorted({r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State" for r in records}):
        reg_count=len({str(r.get("gstin")).strip() for r in records if (r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State")==state and r.get("gstin")})
        wa.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=9); wa.cell(rr,1,f"{state}  •  {reg_count} registration(s)"); wa.cell(rr,1).fill=PatternFill("solid",fgColor=light); wa.cell(rr,1).font=Font(name="Trebuchet MS",size=10,bold=True,color=maroon); rr+=1
        state_recs=sorted([r for r in records if (r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State")==state], key=lambda r:r.get("month") or datetime.max)
        for rec in state_recs:
            for x in rec["sections"]:
                detail=f"Taxable={x['taxable']:.2f}; IGST={x['igst']:.2f}; CGST={x['cgst']:.2f}; SGST={x['sgst']:.2f}; Cess={x['cess']:.2f}"
                vals=[state,month_display(rec["month"]),rec["gstin"],rec["fy"],rec.get("file_name",""),x["page"],x["section"],x["head"],detail]
                for c,v in enumerate(vals,1): wa.cell(rr,c,v)
                rr+=1
            for x in rec["payment"]:
                detail=f"Payable={x['tax_payable']:.2f}; Net={x['net_tax_payable']:.2f}; ITC={x['itc_igst']+x['itc_cgst']+x['itc_sgst']+x['itc_cess']:.2f}; Cash={x['cash_tax']:.2f}"
                vals=[state,month_display(rec["month"]),rec["gstin"],rec["fy"],rec.get("file_name",""),x["page"],x["section"],x["head"],detail]
                for c,v in enumerate(vals,1): wa.cell(rr,c,v)
                rr+=1
        rr += 1
    wa.freeze_panes="A2"; wa.auto_filter.ref=wa.dimensions
    for c,w in enumerate([22,13,20,12,42,8,34,72,80],1): wa.column_dimensions[get_column_letter(c)].width=w

    # ------------------------------------------------------------------
    # Consolidated Summary — STATE-WISE, MONTH-WISE.
    # Multiple GSTINs belonging to the same state are consolidated into
    # ONE row per month (no duplicate Apr/May/etc. rows).
    # ------------------------------------------------------------------
    if "Consolidated Summary" in wb.sheetnames:
        del wb["Consolidated Summary"]
    cs = wb.create_sheet("Consolidated Summary", 0)
    maroon = "98002E"; light = "F3E6EB"
    cs["A1"] = "GSTR-3B CONSOLIDATED SUMMARY — STATE WISE"
    cs.merge_cells("A1:J1")
    cs["A1"].fill = PatternFill("solid", fgColor=maroon)
    cs["A1"].font = Font(name="Trebuchet MS", size=10, bold=True, color="FFFFFF")
    cs["A1"].alignment = Alignment(horizontal="center", vertical="center")
    cs["A2"] = "Grouping: State → Month. Multiple GSTINs in the same state are consolidated into one monthly row. NIL-filed months are retained with zero values."
    cs.merge_cells("A2:J2")
    cs["A2"].font = Font(name="Trebuchet MS", size=10, italic=True, color="666666")

    headers = [
        "State","Month","Filing Status",
        "3.1 Taxable","3.1 IGST","3.1 CGST","3.1 SGST","3.1 Cess",
        "ITC IGST","ITC CGST","ITC SGST","ITC Cess"
    ]
    # Recreate header for the actual 12 columns.
    for c,h in enumerate(headers,1):
        cell=cs.cell(3,c,h)
        cell.fill=PatternFill("solid", fgColor=maroon)
        cell.font=Font(name="Trebuchet MS", size=10, bold=True, color="FFFFFF")
        cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)

    def section_total(rec, sections):
        vals=[0.0,0.0,0.0,0.0,0.0]
        for x in rec.get("sections",[]):
            if x.get("section") in sections:
                vals[0]+=x.get("taxable",0) or 0
                vals[1]+=x.get("igst",0) or 0
                vals[2]+=x.get("cgst",0) or 0
                vals[3]+=x.get("sgst",0) or 0
                vals[4]+=x.get("cess",0) or 0
        return vals

    def itc_total(rec):
        vals=[0.0,0.0,0.0,0.0]
        for x in rec.get("sections",[]):
            if str(x.get("section","" )).startswith("4"):
                vals[0]+=x.get("igst",0) or 0
                vals[1]+=x.get("cgst",0) or 0
                vals[2]+=x.get("sgst",0) or 0
                vals[3]+=x.get("cess",0) or 0
        return vals

    # State -> Month -> records. This is the key consolidation step.
    grouped={}
    for r in records:
        state=r.get("state") or state_from_gstin(r.get("gstin")) or "Unknown State"
        month=r.get("month")
        grouped.setdefault(state, {}).setdefault(month, []).append(r)

    def combined_status(recs):
        statuses={str(r.get("filing_status") or "FILED").upper() for r in recs}
        if "FILED" in statuses:
            return "FILED"
        if "NIL FILED" in statuses or "NIL" in statuses:
            return "NIL FILED"
        return next(iter(statuses), "FILED")

    row=4
    grand=[0.0]*9
    for state in sorted(grouped):
        state_months=grouped[state]
        registration_count=len({r.get("gstin") for recs in state_months.values() for r in recs if r.get("gstin")})
        if row > 4:
            row += 1
        cs.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cs.cell(row,1,f"{state}  •  {registration_count} registration(s)")
        cs.cell(row,1).fill=PatternFill("solid", fgColor=light)
        cs.cell(row,1).font=Font(name="Trebuchet MS", size=10, bold=True, color=maroon)
        row += 1

        subtotal=[0.0]*9
        for month in sorted(state_months, key=lambda x: x or datetime.max):
            recs=state_months[month]
            vals=[0.0]*9
            for rec in recs:
                outvals=section_total(rec,{"3.1"})
                itc=itc_total(rec)
                recvals=outvals[:]+itc
                vals=[a+b for a,b in zip(vals,recvals)]
            subtotal=[a+b for a,b in zip(subtotal,vals)]
            grand=[a+b for a,b in zip(grand,vals)]

            out=[state,month_display(month),combined_status(recs),*vals]
            for c,v in enumerate(out,1):
                cs.cell(row,c,v)
                cs.cell(row,c).font=Font(name="Trebuchet MS",size=10)
            row += 1

        cs.cell(row,1,f"{state} Total")
        for c,v in enumerate(subtotal,4): cs.cell(row,c,v)
        for c in range(1,len(headers)+1):
            cs.cell(row,c).fill=PatternFill("solid",fgColor=light)
            cs.cell(row,c).font=Font(name="Trebuchet MS",size=10,bold=True,color=maroon)
        row += 1

    cs.cell(row,1,"GRAND TOTAL — ALL STATES / GSTINs")
    for c,v in enumerate(grand,4): cs.cell(row,c,v)
    for c in range(1,len(headers)+1):
        cs.cell(row,c).fill=PatternFill("solid",fgColor=maroon)
        cs.cell(row,c).font=Font(name="Trebuchet MS",size=10,bold=True,color="FFFFFF")
    for rr in range(4,row+1):
        for c in range(4,len(headers)+1):
            cs.cell(rr,c).number_format="#,##0.00"
    for c,w in enumerate([25,13,16,18,16,16,16,16,16,16,16,16],1):
        cs.column_dimensions[get_column_letter(c)].width=w
    cs.freeze_panes="A4"
    cs.auto_filter.ref=f"A3:L{row}"

    # Reorder sheets to a logical flow matching the supplied workbook.
    desired=[
        "Consolidated Summary",
        "Read Me",
        "3.1 Outward Details",
        "3.1.1 Supplies u-s 9(5)",
        "3.2 Inter-State Supplies",
        "4. ITC Details",
        "5. Inward Supplies",
        "5.1 Interest & Late Fee",
        "GSTR-3B Detailed",
        "6.1 Payment of Tax",
        "Extraction Audit",
    ]
    wb._sheets=[wb[x] for x in desired if x in wb.sheetnames]

    # Global workbook typography: Trebuchet MS, 10 pt, while preserving
    # bold/italic/color/alignment and all existing template formatting.
    for sh in wb.worksheets:
        for row_cells in sh.iter_rows():
            for cell in row_cells:
                oldf = copy(cell.font)
                cell.font = Font(
                    name="Trebuchet MS", size=10, bold=oldf.bold, italic=oldf.italic,
                    vertAlign=oldf.vertAlign, underline=oldf.underline,
                    strike=oldf.strike, color=oldf.color, charset=oldf.charset,
                    family=oldf.family, scheme=oldf.scheme, outline=oldf.outline,
                    shadow=oldf.shadow, condense=oldf.condense, extend=oldf.extend
                )
    return wb


def main():
    st.set_page_config(
        page_title="Pushpak Kumar | GSTR-3B Analyzer",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    st.markdown("""
    <style>
    .stApp {background:#f7f8fb;}
    .main .block-container {max-width:1180px;padding-top:2rem;padding-bottom:3rem;}
    .hero {background:linear-gradient(135deg,#98002e,#65001f);border-radius:20px;padding:30px 34px;color:#fff;margin-bottom:22px;box-shadow:0 10px 30px rgba(90,0,30,.16);}
    .brand-row {display:flex;align-items:center;gap:12px;margin-bottom:20px;}
    .brand-logo {width:44px;height:44px;border-radius:12px;background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.25);display:flex;align-items:center;justify-content:center;font-size:17px;font-weight:800;}
    .brand-name {font-size:14px;font-weight:750;letter-spacing:.4px;color:#fff;}
    .brand-tagline {font-size:11px;color:rgba(255,255,255,.72);margin-top:2px;}
    .hero h1 {margin:0 0 8px;font-size:31px;font-weight:750;letter-spacing:-.6px;}
    .hero p {margin:0;color:rgba(255,255,255,.88);font-size:14px;}
    .card {background:#fff;border:1px solid #e7e9ef;border-radius:16px;padding:22px;margin-bottom:18px;box-shadow:0 3px 14px rgba(20,20,40,.05);}
    .step {display:flex;align-items:center;gap:8px;flex:1;}
    .step-num {width:28px;height:28px;border-radius:50%;background:#f3e6eb;color:#98002e;display:flex;align-items:center;justify-content:center;font-weight:750;}
    .workflow {display:flex;gap:10px;align-items:center;color:#424752;font-size:13px;}
    .workflow-line {height:1px;background:#dddfe6;flex:1;}
    .section-title {font-size:18px;font-weight:700;color:#242733;margin-bottom:4px;}
    .section-subtitle {color:#707582;font-size:13px;margin-bottom:14px;}
    .status-card {background:#fff;border-left:4px solid #98002e;border-radius:10px;padding:13px 16px;margin:12px 0 16px;box-shadow:0 2px 9px rgba(20,20,40,.04);}
    .brand-chip {display:inline-flex;align-items:center;gap:7px;background:#f3e6eb;color:#98002e;border-radius:999px;padding:6px 11px;font-size:11px;font-weight:750;margin-bottom:11px;}
    .brand-dot {width:6px;height:6px;border-radius:50%;background:#98002e;}
    .footer {text-align:center;color:#858994;font-size:12px;padding-top:22px;}
    .footer-name {color:#98002e;font-weight:750;}
    div[data-testid="stFileUploader"] {background:#fbfbfd;border:1px dashed #c9ccd6;border-radius:13px;padding:8px;}
    .stButton > button,.stDownloadButton > button {border-radius:10px;min-height:44px;font-weight:650;}
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="hero">
      <div class="brand-row">
        <div class="brand-logo">PK</div>
        <div>
          <div class="brand-name">PUSHPAK KUMAR</div>
          <div class="brand-tagline">GSTR-3B Analysis &amp; Reporting</div>
        </div>
      </div>
      <h1>GSTR-3B PDF → Excel Analyzer</h1>
      <p>Automate filed GSTR-3B extraction, detailed mapping and monthly consolidation.</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="card">
      <div class="workflow">
        <div class="step"><span class="step-num">1</span><b>Upload ZIP</b></div>
        <div class="workflow-line"></div>
        <div class="step"><span class="step-num">2</span><b>Analyze PDFs</b></div>
        <div class="workflow-line"></div>
        <div class="step"><span class="step-num">3</span><b>Validate</b></div>
        <div class="workflow-line"></div>
        <div class="step"><span class="step-num">4</span><b>Download Excel</b></div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="card">
      <div class="brand-chip"><span class="brand-dot"></span> PUSHPAK KUMAR • ANALYZER</div>
      <div class="section-title">Upload filed GSTR-3B PDFs</div>
      <div class="section-subtitle">Upload one ZIP containing the monthly filed GSTR-3B PDF copies.</div>
    """, unsafe_allow_html=True)

    zip_file = st.file_uploader("Upload ZIP", type=["zip"], label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

    if not zip_file:
        st.info("Recommended: keep PDFs for the same GSTIN together for clean monthly consolidation.")
        st.markdown('<div class="footer">Designed &amp; developed by <span class="footer-name">Pushpak Kumar</span></div>', unsafe_allow_html=True)
        return

    try:
        with zipfile.ZipFile(zip_file) as z:
            pdf_names = [n for n in z.namelist() if n.lower().endswith(".pdf") and not n.endswith("/")]
    except zipfile.BadZipFile:
        st.error("The uploaded file is not a valid ZIP.")
        return

    if not pdf_names:
        st.error("No PDF files were found in the ZIP.")
        return

    st.markdown(f'<div class="status-card"><b>Ready to analyze</b><br><span style="color:#707582;font-size:13px;">{len(pdf_names)} PDF file(s) &nbsp; • &nbsp; {len(zip_file.getvalue())/(1024*1024):.2f} MB</span></div>', unsafe_allow_html=True)

    if st.button("🔍  Analyze GSTR-3B ZIP", type="primary", use_container_width=True):
        records, errors = [], []
        progress = st.progress(0, text="Starting analysis…")
        status = st.empty()

        with zipfile.ZipFile(zip_file) as z:
            for i, name in enumerate(pdf_names, 1):
                status.markdown(f"**Processing {i} of {len(pdf_names)}:** `{Path(name).name}`")
                try:
                    d = extract_pdf(z.read(name))
                    d["file_name"] = name
                    d["state"] = state_from_gstin(d.get("gstin"))
                    d["filing_status"] = "NIL FILED" if is_nil_record(d) else "FILED"
                    if not d.get("month"):
                        raise ValueError(
                            f"Could not identify Financial Year / Tax Period "
                            f"(detected FY={d.get('fy') or 'blank'}, "
                            f"Period={d.get('tax_period') or 'blank'}). "
                            f"Please ensure this is a filed GSTR-3B PDF."
                        )
                    records.append(d)
                except Exception as e:
                    errors.append((name, str(e)))
                progress.progress(i / len(pdf_names), text=f"Analyzing PDF {i} of {len(pdf_names)}")

        progress.progress(1.0, text="Analysis complete ✓")
        status.empty()

        if errors:
            st.warning(f"{len(errors)} file(s) could not be parsed.")
            with st.expander("View parsing errors"):
                st.dataframe([{"File": n, "Error": e} for n, e in errors], use_container_width=True, hide_index=True)

        if not records:
            st.error("No GSTR-3B PDFs could be successfully parsed.")
            return

        gstins = sorted({r["gstin"] for r in records if r.get("gstin")})
        fys = sorted({r["fy"] for r in records if r.get("fy")})
        a,b,c = st.columns(3)
        a.metric("PDFs Parsed", len(records))
        b.metric("GSTINs Found", len(gstins))
        c.metric("Financial Year(s)", len(fys))

        preview = [{
            "State": r.get("state", state_from_gstin(r.get("gstin"))),
            "Month": month_display(r["month"]),
            "GSTIN": r["gstin"],
            "Filing Status": r.get("filing_status","FILED"),
            "Tax Period": r["tax_period"],
            "Financial Year": r["fy"],
            "File": Path(r["file_name"]).name,
        } for r in sorted(records, key=lambda x: (x.get("state", ""), x["month"] or datetime.max, x.get("gstin", "")))]

        st.markdown('<div class="section-title" style="margin-top:22px;">Extracted periods</div><div class="section-subtitle">Review the detected periods before downloading.</div>', unsafe_allow_html=True)
        st.dataframe(preview, use_container_width=True, hide_index=True)

        build = st.progress(0, text="Preparing Excel workbook…")
        build.progress(30, text="Applying GSTR-3B table mappings…")
        wb = build_workbook(records)
        build.progress(65, text="Building detailed and payment sheets…")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        build.progress(100, text="Excel report ready ✓")

        st.success("GSTR-3B consolidated workbook generated successfully.")
        st.download_button(
            "⬇️  Download GSTR3B_Analyzed.xlsx",
            out.getvalue(),
            file_name="GSTR3B_Analyzed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    st.markdown('<div class="footer">GSTR-3B PDF → Excel Analyzer<br>Designed &amp; developed by <span class="footer-name">Pushpak Kumar</span></div>', unsafe_allow_html=True)


if __name__ == "__main__":
    main()
